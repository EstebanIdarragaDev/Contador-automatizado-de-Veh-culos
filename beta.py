# Importacion de OpenCV
import cv2
import numpy as np
import imutils

# Librerias para crear el excel
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font
import datetime


# ----------------------------------------------------------
cap = cv2.VideoCapture('./cars.mp4')
book = Workbook();
sheet = book.active;
sheet['A1'] = 'FECHA'
sheet['B1'] = 'MOVIMIENTO'
sheet['C1'] = 'TIPO_VEHICULO'
sheet['D1'] = 'VOLUMEN'


fgbg = cv2.bgsegm.createBackgroundSubtractorMOG()
kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))
car_counter1 = 0
car_counter2 = 0



while True:
    ret, frame = cap.read()
    if ret == False: break
    frame = imutils.resize(frame, width=640) #tamaño de la ventana

    # Especificamos los puntos extremos del área a analizar
    # Area a analizar
    area_pts = np.array([[200, 10], [frame.shape[1]-90, 10], [frame.shape[1]-90, 180], [200, 180]])
    area_pts2 = np.array([[200, 190], [frame.shape[1]-90, 190], [frame.shape[1]-90, 350], [200, 350]])


    # Con ayuda de una imagen auxiliar, determinamos el área
    # sobre la cual actuará el detector de movimiento
    imAux = np.zeros(shape=(frame.shape[:2]), dtype=np.uint8)
    imAux = cv2.drawContours(imAux, [area_pts], -1, (255), -1)
    image_area = cv2.bitwise_and(frame, frame, mask=imAux) 
    # Obtendremos la imagen binaria donde la región en blanco representa
    # la existencia de movimiento
    fgmask = fgbg.apply(image_area)
    fgmask = cv2.morphologyEx(fgmask, cv2.MORPH_OPEN, kernel)
    fgmask = cv2.morphologyEx(fgmask, cv2.MORPH_CLOSE, kernel)
    fgmask = cv2.dilate(fgmask, None, iterations=5)   



    # ------------------------------------------------------------------------- 
    imAux2 = np.zeros(shape=(frame.shape[:2]), dtype=np.uint8)
    imAux2 = cv2.drawContours(imAux2, [area_pts2], -1, (255), -1)
    image_area2 = cv2.bitwise_and(frame, frame, mask=imAux2) 
    # Obtendremos la imagen binaria donde la región en blanco representa
    # la existencia de movimiento
    fgmask2 = fgbg.apply(image_area2)
    fgmask2 = cv2.morphologyEx(fgmask2, cv2.MORPH_OPEN, kernel)
    fgmask2 = cv2.morphologyEx(fgmask2, cv2.MORPH_CLOSE, kernel)
    fgmask2 = cv2.dilate(fgmask2, None, iterations=5)   



    # Encontramos los contornos presentes de fgmask, para luego basándonos
    # en su área poder determinar si existe movimiento (autos)
    cnts = cv2.findContours(fgmask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[0]
    for cnt in cnts:
        if cv2.contourArea(cnt) > 1500:
            x, y, w, h = cv2.boundingRect(cnt)
            cv2.rectangle(frame, (x,y), (x+w,y+h), (0,255,255), 1)  
            # Si el auto ha cruzado entre 440 y 460 abierto, se incrementará
        # en 1 el contador de autos
            if 440 < (x + w) < 460:
                car_counter1 = car_counter1 + 1
                print(h);
                if h>50:print('camion')
                cv2.line(frame, (380, 10), (380, 180), (0, 255, 0), 3)
                sheet[f'A{sheet.max_row+1}'] = datetime.datetime.now()
                sheet[f'B{sheet.max_row}'] = 1
                sheet[f'C{sheet.max_row}'] = 'AUTOMOVIL'
                sheet[f'D{sheet.max_row}'] = 1



    cnts2 = cv2.findContours(fgmask2, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[0]
    for cnt in cnts2:
        if cv2.contourArea(cnt) > 1500:
            x, y, w, h = cv2.boundingRect(cnt)
            cv2.rectangle(frame, (x,y), (x+w,y+h), (0,255,255), 1)  
            # Si el auto ha cruzado entre 440 y 460 abierto, se incrementará
            # en 1 el contador de autos
            if 440 < (x + w) < 460:
                car_counter2 = car_counter2 + 1
                cv2.line(frame, (380, 190), (380, 350), (0, 255, 0), 3)
                sheet[f'A{sheet.max_row+1}'] = datetime.datetime.now()
                sheet[f'B{sheet.max_row}'] = 2
                sheet[f'C{sheet.max_row}'] = 'AUTOMOVIL'
                sheet[f'D{sheet.max_row}'] = 1
                    
                  
    # Visualización del conteo de autos
    cv2.drawContours(frame, [area_pts], -1, (255, 0, 255), 2)
    cv2.drawContours(frame, [area_pts2], -1, (255, 0, 0), 2)
    cv2.line(frame, (380, 10), (380, 180), (0, 255, 255), 1)
    cv2.line(frame, (380, 190), (380, 350), (0, 255, 255), 1)



    # cv2.rectangle(frame, (frame.shape[1]-70, 215), (frame.shape[1]-5, 270), (0, 255, 0), 2)
    # cv2.putText(frame, str(car_counter), (frame.shape[1]-55, 250), cv2.FONT_HERSHEY_SIMPLEX, 1.2, 
    #     (0,255,0), 2)
    cv2.imshow('frame', frame)


    k = cv2.waitKey(1) & 0xFF
    if k == 27:
        print(car_counter1,car_counter2);
        break
    
print(car_counter1,car_counter2);
cap.release()
cv2.destroyAllWindows()
book.save(f'Aforo {datetime.datetime.now().date()}.xlsx');